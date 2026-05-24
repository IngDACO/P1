"""
Exportar e importar la matriz SURVEY como Excel.

Hojas:
  - SURVEY: la matriz editable
  - INFO  : parámetros del proyecto (PDF + usuario)
  - CONFIG: condiciones (omega, pared limitante, controlador, paradas)

import_survey_excel devuelve un dict con:
  - df:     DataFrame con la matriz survey
  - info:   dict con parámetros numéricos (PDF + usuario)
  - config: dict con condiciones (puede estar vacío si Excel antiguo)
"""
import io
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

SURVEY_COLS = ["WR", "FR", "OR", "WL", "FL", "OL"]

# Condiciones del proyecto que se guardan en hoja CONFIG
CONFIG_KEYS = [
    "OMEGA_SIDE", "WALL_LIMITING", "WALL_STOP", "WALL_SIDE",
    "OFFSET_SIDE", "CTRL_IN_FRAME", "CTRL_SIDE", "NS",
]

_HEADER_FILL = PatternFill("solid", fgColor="1A3A5C")
_HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
_NORMAL_FONT = Font(name="Calibri", size=10)
_CENTER      = Alignment(horizontal="center", vertical="center")
_BORDER      = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _write_styled_header(ws, row, col, text, fill=_HEADER_FILL):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = _HEADER_FONT
    cell.fill      = fill
    cell.alignment = _CENTER
    cell.border    = _BORDER
    return cell


def export_survey_excel(df: pd.DataFrame, project_info: dict = None,
                         project_config: dict = None) -> bytes:
    """
    Genera un Excel con hojas SURVEY, INFO y CONFIG.
    project_info:   dict de parámetros numéricos (PDF + usuario)
    project_config: dict de condiciones (omega side, wall limiting, etc.)
    """
    wb = openpyxl.Workbook()

    # ── Hoja SURVEY ──
    ws = wb.active
    ws.title = "SURVEY"

    # Título
    ws.merge_cells("A1:G1")
    ws["A1"] = "SURVEY ANALYZER — Matriz de medidas en campo (mm)"
    ws["A1"].font      = Font(color="FFFFFF", bold=True, name="Calibri", size=12)
    ws["A1"].fill      = _HEADER_FILL
    ws["A1"].alignment = _CENTER

    # Encabezados
    headers = ["Parada"] + SURVEY_COLS
    for col_idx, h in enumerate(headers, start=1):
        _write_styled_header(ws, 2, col_idx, h)

    # Datos
    for row_idx, (_, row) in enumerate(df.iterrows(), start=3):
        cell = ws.cell(row=row_idx, column=1, value=row_idx - 2)
        cell.font      = Font(bold=True, name="Calibri", size=10)
        cell.alignment = _CENTER
        cell.border    = _BORDER
        for col_idx, col in enumerate(SURVEY_COLS, start=2):
            c = ws.cell(row=row_idx, column=col_idx, value=float(row[col]))
            c.font      = _NORMAL_FONT
            c.alignment = _CENTER
            c.border    = _BORDER

    ws.column_dimensions["A"].width = 10
    for col_letter in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = 14

    # ── Hoja INFO (parámetros numéricos) ──
    if project_info:
        _write_info_sheet(wb, "INFO", "Parámetros del proyecto", project_info)

    # ── Hoja CONFIG (condiciones del proyecto) ──
    if project_config:
        _write_info_sheet(wb, "CONFIG", "Condiciones del proyecto", project_config)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _write_info_sheet(wb, sheet_name: str, title: str, data: dict) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.merge_cells("A1:B1")
    ws["A1"]           = title
    ws["A1"].font      = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    ws["A1"].fill      = _HEADER_FILL
    ws["A1"].alignment = _CENTER

    _write_styled_header(ws, 2, 1, "Parámetro", fill=PatternFill("solid", fgColor="2E6DA4"))
    _write_styled_header(ws, 2, 2, "Valor",     fill=PatternFill("solid", fgColor="2E6DA4"))

    row_idx = 3
    for k, v in data.items():
        if v is None:
            continue
        if isinstance(v, bool):
            disp = "True" if v else "False"
        elif isinstance(v, float):
            disp = str(round(v, 3))
        else:
            disp = str(v)
        c1 = ws.cell(row=row_idx, column=1, value=str(k))
        c2 = ws.cell(row=row_idx, column=2, value=disp)
        c1.font = _NORMAL_FONT; c1.border = _BORDER
        c2.font = _NORMAL_FONT; c2.border = _BORDER
        row_idx += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20


# ── Importar ─────────────────────────────────────────────────
def _parse_config_value(key: str, raw: str):
    """Convierte el valor crudo (string) al tipo apropiado."""
    if raw is None or str(raw).strip() == "":
        return None
    s = str(raw).strip()
    if key in ("WALL_LIMITING", "CTRL_IN_FRAME"):
        return s.lower() in ("true", "1", "yes", "y", "sí", "si")
    if key in ("WALL_STOP", "NS"):
        try:    return int(float(s))
        except: return None
    if key in ("OMEGA_SIDE", "WALL_SIDE", "OFFSET_SIDE", "CTRL_SIDE"):
        return s if s in ("R", "L") else None
    return s


def import_survey_excel(file) -> dict:
    """
    Lee un Excel generado por export_survey_excel.

    Returns dict con keys:
      - df:     DataFrame con la matriz survey
      - info:   dict de parámetros numéricos (puede estar vacío)
      - config: dict de condiciones (puede estar vacío)
    """
    result = {"df": None, "info": {}, "config": {}}

    try:
        df = pd.read_excel(file, sheet_name="SURVEY", header=1, usecols="A:G")
    except Exception as e:
        raise ValueError(f"No se pudo leer la hoja SURVEY: {e}")

    expected_cols = ["Parada"] + SURVEY_COLS
    if len(df.columns) != len(expected_cols):
        raise ValueError(f"La hoja SURVEY debe tener {len(expected_cols)} columnas (encontradas: {len(df.columns)})")
    df.columns = expected_cols
    df = df.drop(columns=["Parada"])
    df = df.apply(pd.to_numeric, errors="coerce").fillna(0.0)
    result["df"] = df

    # ── INFO (opcional) ──
    try:
        info_df = pd.read_excel(file, sheet_name="INFO", header=1)
        for _, row in info_df.iterrows():
            k = str(row.iloc[0]).strip() if len(row) > 0 else ""
            v = row.iloc[1] if len(row) > 1 else None
            if k and pd.notna(v):
                try:    result["info"][k] = float(v)
                except: pass
    except Exception:
        pass

    # ── CONFIG (opcional) ──
    try:
        cfg_df = pd.read_excel(file, sheet_name="CONFIG", header=1)
        for _, row in cfg_df.iterrows():
            k = str(row.iloc[0]).strip() if len(row) > 0 else ""
            v = row.iloc[1] if len(row) > 1 else None
            if k and pd.notna(v) and k in CONFIG_KEYS:
                parsed = _parse_config_value(k, v)
                if parsed is not None:
                    result["config"][k] = parsed
    except Exception:
        pass

    return result
